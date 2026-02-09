#!/usr/bin/env python3
"""
HA Master Context Excel Export v3.0
Enhanced with proper headers, AutoFilter, Excel formulas, and LLM-optimized outputs
"""

import sys
import os
import re
import csv
import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Installing...", file=sys.stderr)
    os.system("pip3 install openpyxl --break-system-packages")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# ============================================================================
# SHEET DEFINITIONS - Single source of truth for all columns
# ============================================================================

SHEET_DEFINITIONS = {
    'Entities': ['entity_id', 'platform', 'area_id', 'device_id', 'disabled_by', 'original_name', 'unique_id', 'domain'],
    'Devices': ['id', 'name', 'manufacturer', 'model', 'sw_version', 'area_id', 'identifiers', 'device_type'],
    'Areas': ['id', 'name', 'floor_id', 'aliases'],
    'Automations': ['id', 'alias', 'description', 'trigger_types', 'entity_count', 'file_location', 'mode'],
    'Integrations': ['entry_id', 'domain', 'title', 'source', 'version', 'data_keys'],
    'Action Items': ['entity_id', 'domain', 'friendly_name', 'platform', 'issue_type', 'priority', 'recommendation'],
    'HAC Learnings': ['date', 'filename', 'project', 'file_path', 'size_kb']
}

# ============================================================================
# STYLING FUNCTIONS
# ============================================================================

def apply_header_style(ws, row=1):
    """Apply Office 365 style to header row"""
    header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

def auto_size_columns(ws, max_width=50):
    """Auto-size columns based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def apply_autofilter(ws):
    """Apply AutoFilter to data range"""
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

# ============================================================================
# DATA WRITING FUNCTIONS
# ============================================================================

def _write_data_sheet(wb, csv_path, sheet_name, headers):
    """
    Unified function to write data sheet with consistent formatting
    
    Args:
        wb: Workbook object
        csv_path: Path to CSV file
        sheet_name: Name for the sheet
        headers: List of column headers from SHEET_DEFINITIONS
    
    Returns:
        Worksheet object or None if file not found
    """
    if not os.path.exists(csv_path):
        print(f"WARNING: {csv_path} not found, skipping {sheet_name}...", file=sys.stderr)
        return None
    
    ws = wb.create_sheet(sheet_name)
    
    # Write headers first
    ws.append(headers)
    
    # Write data rows
    with open(csv_path, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
    
    # Apply consistent formatting
    apply_header_style(ws, row=1)
    apply_autofilter(ws)
    auto_size_columns(ws)
    ws.freeze_panes = 'A2'
    
    return ws

def create_dashboard_sheet(wb, temp_dir):
    """Create dashboard with Excel formulas for live counts"""
    ws = wb.create_sheet("Dashboard", 0)
    
    # Title
    ws['A1'] = "Home Assistant Master Context v3.0"
    ws['A1'].font = Font(size=16, bold=True, color="0078D4")
    
    # Metadata
    ws['A3'] = "Generated:"
    ws['B3'] = os.path.basename(temp_dir).replace('temp_', '')
    ws['A3'].font = Font(bold=True)
    
    # Summary Statistics (using formulas!)
    ws['A5'] = "SUMMARY STATISTICS"
    ws['A5'].font = Font(size=14, bold=True, color="0078D4")
    
    ws['A6'] = "Total Entities:"
    ws['B6'] = '=COUNTA(Entities!A:A)-1'  # Count non-empty cells minus header
    
    ws['A7'] = "Total Devices:"
    ws['B7'] = '=COUNTA(Devices!A:A)-1'
    
    ws['A8'] = "Total Areas:"
    ws['B8'] = '=COUNTA(Areas!A:A)-1'
    
    ws['A9'] = "Total Automations:"
    ws['B9'] = '=COUNTA(Automations!A:A)-1'
    
    ws['A10'] = "Total Integrations:"
    ws['B10'] = '=COUNTA(Integrations!A:A)-1'
    
    # Action Items Breakdown
    ws['A12'] = "ACTION ITEMS"
    ws['A12'].font = Font(size=14, bold=True, color="C43900")
    
    ws['A13'] = "Total Items:"
    ws['B13'] = '=COUNTA("Action Items"!A:A)-1'
    ws['B13'].font = Font(bold=True, color="C43900")
    
    ws['A14'] = "HIGH Priority:"
    ws['B14'] = '=COUNTIF("Action Items"!F:F,"HIGH")'
    ws['B14'].font = Font(color="C43900")
    
    ws['A15'] = "MEDIUM Priority:"
    ws['B15'] = '=COUNTIF("Action Items"!F:F,"MEDIUM")'
    
    ws['A16'] = "LOW Priority:"
    ws['B16'] = '=COUNTIF("Action Items"!F:F,"LOW")'
    
    # Entity Health
    ws['A18'] = "ENTITY HEALTH"
    ws['A18'].font = Font(size=14, bold=True, color="107C10")
    
    ws['A19'] = "Entities Without Areas:"
    ws['B19'] = '=COUNTIF(Entities!C:C,"UNASSIGNED")'
    
    ws['A20'] = "Entities Without Devices:"
    ws['B20'] = '=COUNTIF(Entities!D:D,"NO_DEVICE")'
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

# ============================================================================
# PRIVACY FUNCTIONS
# ============================================================================

def redact_sensitive_data(value):
    """Redact sensitive information for AI_SAFE version"""
    if not isinstance(value, str):
        return value
    
    # Redact IPs
    value = re.sub(r'\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b', '[IP_REDACTED]', value)
    
    # Redact MACs
    value = re.sub(r'([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})', '[MAC_REDACTED]', value)
    
    # Redact GPS
    value = re.sub(r'-?\d+\.\d{6,}, ?-?\d+\.\d{6,}', '[GPS_REDACTED]', value)
    
    # Redact personal names
    personal_patterns = [
        r'[Jj]ohn[s]?[_\s]', r'[Mm]ichelle[s]?[_\s]',
        r'[Aa]laina[s]?[_\s]', r'[Ee]lla[s]?[_\s]',
        r'[Jj]arrett[s]?[_\s]', r'[Oo]wen[s]?[_\s]'
    ]
    
    for pattern in personal_patterns:
        value = re.sub(pattern, 'User_', value)
    
    return value

def create_safe_version(wb_full, temp_dir):
    """Create redacted AI_SAFE version"""
    wb_safe = Workbook()
    wb_safe.remove(wb_safe.active)
    
    for sheet_name in wb_full.sheetnames:
        ws_full = wb_full[sheet_name]
        ws_safe = wb_safe.create_sheet(sheet_name)
        
        for row in ws_full.iter_rows():
            safe_row = []
            for cell in row:
                safe_value = redact_sensitive_data(cell.value)
                safe_row.append(safe_value)
            
            ws_safe.append(safe_row)
        
        # Copy formatting
        if sheet_name != "Dashboard":
            apply_header_style(ws_safe)
            apply_autofilter(ws_safe)
            auto_size_columns(ws_safe)
            ws_safe.freeze_panes = 'A2'
    
    return wb_safe

# ============================================================================
# LLM-OPTIMIZED OUTPUT FUNCTIONS
# ============================================================================

def generate_llm_json(temp_dir, output_dir, timestamp):
    """Generate JSON summary for LLM context windows"""
    
    summary = {
        "generated": timestamp,
        "version": "3.0",
        "statistics": {},
        "areas": [],
        "action_items_by_priority": {"HIGH": [], "MEDIUM": [], "LOW": []},
        "recent_learnings": []
    }
    
    # Load areas
    areas_path = os.path.join(temp_dir, 'areas.csv')
    if os.path.exists(areas_path):
        with open(areas_path, 'r') as f:
            reader = csv.DictReader(f, fieldnames=SHEET_DEFINITIONS['Areas'])
            summary['areas'] = [row for row in reader]
    
    # Sample action items (5 per priority for token efficiency)
    action_items_path = os.path.join(temp_dir, 'action_items.csv')
    if os.path.exists(action_items_path):
        with open(action_items_path, 'r') as f:
            reader = csv.DictReader(f, fieldnames=SHEET_DEFINITIONS['Action Items'])
            next(reader)  # Skip header if present
            
            for row in reader:
                priority = row.get('priority', 'MEDIUM')
                if len(summary['action_items_by_priority'].get(priority, [])) < 5:
                    summary['action_items_by_priority'][priority].append(row)
    
    # Recent learnings (last 10)
    learnings_path = os.path.join(temp_dir, 'learnings.csv')
    if os.path.exists(learnings_path):
        with open(learnings_path, 'r') as f:
            reader = csv.DictReader(f, fieldnames=SHEET_DEFINITIONS['HAC Learnings'])
            learnings = list(reader)
            summary['recent_learnings'] = learnings[-10:]
    
    # Statistics
    summary['statistics'] = {
        'total_entities': sum(1 for _ in open(os.path.join(temp_dir, 'entities.csv'))) - 1,
        'total_devices': sum(1 for _ in open(os.path.join(temp_dir, 'devices.csv'))) - 1,
        'total_areas': len(summary['areas']),
        'total_action_items': sum(len(items) for items in summary['action_items_by_priority'].values())
    }
    
    # Write JSON
    json_path = os.path.join(output_dir, f"HA_Master_Context_{timestamp}_LLM.json")
    with open(json_path, 'w') as f:
        json.dump(summary, f, indent=2)
    
    return json_path

# ============================================================================
# MAIN EXPORT FUNCTION
# ============================================================================

def main():
    if len(sys.argv) < 4:
        print("Usage: merge_to_excel.py <temp_dir> <output_dir> <timestamp>", file=sys.stderr)
        sys.exit(1)
    
    temp_dir = sys.argv[1]
    output_dir = sys.argv[2]
    timestamp = sys.argv[3]
    
    # Create FULL workbook
    wb_full = Workbook()
    wb_full.remove(wb_full.active)
    
    # Create Dashboard with formulas
    create_dashboard_sheet(wb_full, temp_dir)
    
    # Import all data sheets with proper headers
    sheet_order = [
        ('action_items.csv', 'Action Items'),
        ('entities.csv', 'Entities'),
        ('devices.csv', 'Devices'),
        ('areas.csv', 'Areas'),
        ('automations.csv', 'Automations'),
        ('integrations.csv', 'Integrations'),
        ('learnings.csv', 'HAC Learnings')
    ]
    
    for csv_file, sheet_name in sheet_order:
        csv_path = os.path.join(temp_dir, csv_file)
        headers = SHEET_DEFINITIONS.get(sheet_name, [])
        _write_data_sheet(wb_full, csv_path, sheet_name, headers)
    
    # Save FULL version
    full_path = os.path.join(output_dir, f"HA_Master_Context_{timestamp}_FULL.xlsx")
    wb_full.save(full_path)
    print(f"✅ Created: {full_path}")
    
    # Create AI_SAFE version
    wb_safe = create_safe_version(wb_full, temp_dir)
    safe_path = os.path.join(output_dir, f"HA_Master_Context_{timestamp}_AI_SAFE.xlsx")
    wb_safe.save(safe_path)
    print(f"✅ Created: {safe_path}")
    
    # Generate LLM-optimized JSON
    json_path = generate_llm_json(temp_dir, output_dir, timestamp)
    print(f"✅ Created LLM JSON: {json_path}")

if __name__ == '__main__':
    main()
